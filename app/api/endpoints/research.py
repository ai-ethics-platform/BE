"""
연구 데이터 분석 API
연구진이 실험 데이터를 조회하고 분석할 수 있는 엔드포인트
"""
from typing import Any, List, Optional
from datetime import datetime
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, RedirectResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, func, select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.deps import get_db
from app.models import (
    User, Room, RoomParticipant, RoundChoice, ConsensusChoice,
    VoiceSession, VoiceParticipant, VoiceRecording
)
from app.schemas.research import (
    ExperimentDataResponse,
    RoomDetailResponse,
    DataStatisticsResponse,
    DeleteTestDataRequest,
    DeleteTestDataResponse,
    UserDataExport,
    RoomDataExport,
    VoiceRecordingsResponse,
    VoiceRecordingItem
)


router = APIRouter()


@router.get("/dashboard")
async def research_dashboard():
    """
    연구 데이터 조회 대시보드로 리다이렉트
    """
    return RedirectResponse(url="/static/research_dashboard.html")


@router.get("/experiments/summary", response_model=DataStatisticsResponse)
async def get_experiment_summary(
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    전체 실험 데이터 통계 요약
    - 총 사용자 수
    - 총 게임 세션 수
    - 총 음성 녹음 수
    - 완료된 게임 수
    """
    # AsyncSession 방식으로 수정
    total_users = (await db.execute(select(func.count()).select_from(User).where(User.is_guest == False))).scalar()
    total_rooms = (await db.execute(select(func.count()).select_from(Room))).scalar()
    total_started_rooms = (await db.execute(select(func.count()).select_from(Room).where(Room.is_started == True))).scalar()
    total_voice_recordings = (await db.execute(select(func.count()).select_from(VoiceRecording))).scalar()
    
    # 라운드 선택 통계
    total_round_choices = (await db.execute(select(func.count()).select_from(RoundChoice))).scalar()
    total_consensus_choices = (await db.execute(select(func.count()).select_from(ConsensusChoice))).scalar()
    
    # 동의 통계
    users_with_consent = (await db.execute(
        select(func.count()).select_from(User).where(
            and_(User.data_consent == True, User.voice_consent == True, User.is_guest == False)
        )
    )).scalar()
    
    return {
        "total_users": total_users,
        "users_with_full_consent": users_with_consent,
        "total_rooms": total_rooms,
        "total_started_rooms": total_started_rooms,
        "total_voice_recordings": total_voice_recordings,
        "total_round_choices": total_round_choices,
        "total_consensus_choices": total_consensus_choices,
        "generated_at": datetime.utcnow()
    }


@router.get("/experiments/export", response_model=ExperimentDataResponse)
async def export_experiment_data(
    started_only: bool = Query(True, description="완료된 게임만 포함"),
    with_consent_only: bool = Query(True, description="동의한 사용자만 포함"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    전체 실험 데이터 export
    - 사용자 정보 (동의 여부에 따라 필터링)
    - 게임 세션 정보
    - 라운드별 선택 데이터
    - 음성 녹음 데이터
    """
    # Room 쿼리 구성 (AsyncSession)
    rooms_query = select(Room)
    if started_only:
        rooms_query = rooms_query.where(Room.is_started == True)
    
    rooms_query = rooms_query.offset(skip).limit(limit)
    result = await db.execute(rooms_query)
    rooms = result.scalars().all()
    
    room_data_list = []
    for room in rooms:
        # 참가자 조회
        participants_result = await db.execute(
            select(RoomParticipant).where(RoomParticipant.room_id == room.id)
        )
        participants = participants_result.scalars().all()
        
        participants_data = []
        for participant in participants:
            user_data = None
            if participant.user_id:
                # User 조회
                user_result = await db.execute(
                    select(User).where(User.id == participant.user_id)
                )
                user = user_result.scalar_one_or_none()
                
                if user:
                    # 동의 체크
                    if with_consent_only and not (user.data_consent and user.voice_consent):
                        continue
                    
                    user_data = {
                        "user_id": user.id,
                        "username": user.username,
                        "email": user.email,
                        "birthdate": user.birthdate,
                        "gender": user.gender,
                        "education_level": user.education_level,
                        "major": user.major,
                        "data_consent": user.data_consent,
                        "voice_consent": user.voice_consent,
                        "created_at": user.created_at
                    }
            
            # 참가자의 라운드별 선택
            round_choices_result = await db.execute(
                select(RoundChoice)
                .where(RoundChoice.participant_id == participant.id)
                .order_by(RoundChoice.round_number)
            )
            round_choices = round_choices_result.scalars().all()
            
            participants_data.append({
                "participant_id": participant.id,
                "nickname": participant.nickname,
                "role_id": participant.role_id,
                "is_host": participant.is_host,
                "user_data": user_data,
                "round_choices": [
                    {
                        "round_number": rc.round_number,
                        "choice": rc.choice,
                        "subtopic": rc.subtopic,
                        "confidence": rc.confidence,
                        "created_at": rc.created_at
                    }
                    for rc in round_choices
                ]
            })
        
        # 합의 선택 데이터
        consensus_result = await db.execute(
            select(ConsensusChoice)
            .where(ConsensusChoice.room_id == room.id)
            .order_by(ConsensusChoice.round_number)
        )
        consensus_choices = consensus_result.scalars().all()
        
        # 음성 세션 데이터
        voice_sessions_result = await db.execute(
            select(VoiceSession).where(VoiceSession.room_id == room.id)
        )
        voice_sessions = voice_sessions_result.scalars().all()
        
        voice_sessions_data = []
        for vs in voice_sessions:
            # 음성 녹음 파일
            recordings_result = await db.execute(
                select(VoiceRecording).where(VoiceRecording.voice_session_id == vs.id)
            )
            voice_recordings = recordings_result.scalars().all()
            
            voice_sessions_data.append({
                "session_id": vs.session_id,
                "started_at": vs.started_at,
                "ended_at": vs.ended_at,
                "is_active": vs.is_active,
                "recordings": [
                    {
                        "id": vr.id,
                        "user_id": vr.user_id,
                        "guest_id": vr.guest_id,
                        "file_path": vr.file_path,
                        "file_size": vr.file_size,
                        "duration": vr.duration,
                        "created_at": vr.created_at,
                        "is_processed": vr.is_processed
                    }
                    for vr in voice_recordings
                ]
            })
        
        room_data_list.append({
            "room_id": room.id,
            "room_code": room.room_code,
            "title": room.title,
            "topic": room.topic,
            "ai_type": room.ai_type,
            "ai_name": room.ai_name,
            "is_started": room.is_started,
            "start_time": room.start_time,
            "created_at": room.created_at,
            "participants": participants_data,
            "consensus_choices": [
                {
                    "round_number": cc.round_number,
                    "choice": cc.choice,
                    "subtopic": cc.subtopic,
                    "confidence": cc.confidence,
                    "created_at": cc.created_at
                }
                for cc in consensus_choices
            ],
            "voice_sessions": voice_sessions_data
        })
    
    # Total count
    total_count_query = select(func.count()).select_from(Room)
    if started_only:
        total_count_query = total_count_query.where(Room.is_started == True)
    total_count = (await db.execute(total_count_query)).scalar()
    
    return {
        "rooms": room_data_list,
        "total_count": total_count,
        "page": skip // limit + 1 if limit > 0 else 1,
        "page_size": limit,
        "exported_at": datetime.utcnow()
    }


@router.get("/experiments/rooms/{room_id}", response_model=RoomDetailResponse)
async def get_room_detail(
    room_id: int,
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    특정 room의 상세 데이터 조회
    - 모든 참가자 정보
    - 라운드별 개인 선택
    - 라운드별 합의 선택
    - 음성 녹음 파일 정보
    """
    # Room 조회
    room_result = await db.execute(select(Room).where(Room.id == room_id))
    room = room_result.scalar_one_or_none()
    
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    
    # 참가자 조회
    participants_result = await db.execute(
        select(RoomParticipant).where(RoomParticipant.room_id == room_id)
    )
    participants = participants_result.scalars().all()
    
    # 참가자 상세 정보
    participants_detail = []
    for participant in participants:
        user_info = None
        if participant.user_id:
            user_result = await db.execute(select(User).where(User.id == participant.user_id))
            user = user_result.scalar_one_or_none()
            if user:
                user_info = {
                    "user_id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "birthdate": user.birthdate,
                    "gender": user.gender,
                    "education_level": user.education_level,
                    "major": user.major,
                    "data_consent": user.data_consent,
                    "voice_consent": user.voice_consent
                }
        
        # 라운드 선택
        round_choices_result = await db.execute(
            select(RoundChoice)
            .where(RoundChoice.participant_id == participant.id)
            .order_by(RoundChoice.round_number)
        )
        round_choices = round_choices_result.scalars().all()
        
        participants_detail.append({
            "participant_id": participant.id,
            "nickname": participant.nickname,
            "role_id": participant.role_id,
            "is_host": participant.is_host,
            "joined_at": participant.joined_at,
            "user_info": user_info,
            "round_choices": [
                {
                    "round_number": rc.round_number,
                    "choice": rc.choice,
                    "subtopic": rc.subtopic,
                    "confidence": rc.confidence,
                    "created_at": rc.created_at
                }
                for rc in round_choices
            ]
        })
    
    # 합의 선택
    consensus_result = await db.execute(
        select(ConsensusChoice)
        .where(ConsensusChoice.room_id == room_id)
        .order_by(ConsensusChoice.round_number)
    )
    consensus_choices = consensus_result.scalars().all()
    
    # 음성 세션
    voice_sessions_result = await db.execute(
        select(VoiceSession).where(VoiceSession.room_id == room_id)
    )
    voice_sessions_list = voice_sessions_result.scalars().all()
    
    voice_sessions = []
    for vs in voice_sessions_list:
        recordings_result = await db.execute(
            select(VoiceRecording).where(VoiceRecording.voice_session_id == vs.id)
        )
        recordings = recordings_result.scalars().all()
        
        voice_sessions.append({
            "session_id": vs.session_id,
            "started_at": vs.started_at,
            "ended_at": vs.ended_at,
            "recordings": [
                {
                    "id": rec.id,
                    "user_id": rec.user_id,
                    "guest_id": rec.guest_id,
                    "file_path": rec.file_path,
                    "file_size": rec.file_size,
                    "duration": rec.duration,
                    "created_at": rec.created_at
                }
                for rec in recordings
            ]
        })
    
    return {
        "room_id": room.id,
        "room_code": room.room_code,
        "title": room.title,
        "topic": room.topic,
        "ai_type": room.ai_type,
        "ai_name": room.ai_name,
        "is_started": room.is_started,
        "start_time": room.start_time,
        "created_at": room.created_at,
        "participants": participants_detail,
        "consensus_choices": [
            {
                "round_number": cc.round_number,
                "choice": cc.choice,
                "subtopic": cc.subtopic,
                "confidence": cc.confidence,
                "created_at": cc.created_at
            }
            for cc in consensus_choices
        ],
        "voice_sessions": voice_sessions
    }


@router.get("/experiments/users/{user_id}")
async def get_user_experiment_data(
    user_id: int,
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    특정 사용자의 모든 실험 참여 데이터
    """
    # User 조회
    user_result = await db.execute(select(User).where(User.id == user_id))
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # 참여한 방들
    participations_result = await db.execute(
        select(RoomParticipant).where(RoomParticipant.user_id == user_id)
    )
    participations = participations_result.scalars().all()
    
    rooms_participated = []
    for participation in participations:
        # Room 조회
        room_result = await db.execute(select(Room).where(Room.id == participation.room_id))
        room = room_result.scalar_one_or_none()
        
        if not room:
            continue
        
        # 해당 방에서의 선택들
        choices_result = await db.execute(
            select(RoundChoice)
            .where(RoundChoice.participant_id == participation.id)
            .order_by(RoundChoice.round_number)
        )
        choices = choices_result.scalars().all()
        
        rooms_participated.append({
            "room_id": room.id,
            "room_code": room.room_code,
            "topic": room.topic,
            "role_id": participation.role_id,
            "nickname": participation.nickname,
            "joined_at": participation.joined_at,
            "choices": [
                {
                    "round_number": c.round_number,
                    "choice": c.choice,
                    "subtopic": c.subtopic,
                    "confidence": c.confidence
                }
                for c in choices
            ]
        })
    
    # 음성 녹음들
    recordings_result = await db.execute(
        select(VoiceRecording).where(VoiceRecording.user_id == user_id)
    )
    voice_recordings = recordings_result.scalars().all()
    
    return {
        "user_id": user.id,
        "username": user.username,
        "email": user.email,
        "birthdate": user.birthdate,
        "gender": user.gender,
        "education_level": user.education_level,
        "major": user.major,
        "data_consent": user.data_consent,
        "voice_consent": user.voice_consent,
        "created_at": user.created_at,
        "rooms_participated": rooms_participated,
        "voice_recordings": [
            {
                "id": vr.id,
                "file_path": vr.file_path,
                "file_size": vr.file_size,
                "duration": vr.duration,
                "created_at": vr.created_at
            }
            for vr in voice_recordings
        ]
    }


@router.post("/experiments/cleanup", response_model=DeleteTestDataResponse)
async def delete_test_data(
    request: DeleteTestDataRequest,
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    테스트 데이터 삭제
    - room_ids: 삭제할 room ID 리스트
    - user_ids: 삭제할 user ID 리스트 (선택사항)
    - delete_voice_files: 음성 파일도 삭제할지 여부
    """
    deleted_rooms = 0
    deleted_users = 0
    deleted_voice_recordings = 0
    
    # Room 삭제 (관련 데이터 cascade 삭제)
    if request.room_ids:
        for room_id in request.room_ids:
            room_result = await db.execute(select(Room).where(Room.id == room_id))
            room = room_result.scalar_one_or_none()
            if room:
                # 관련 음성 녹음 파일 경로 수집
                if request.delete_voice_files:
                    voice_sessions_result = await db.execute(
                        select(VoiceSession).where(VoiceSession.room_id == room_id)
                    )
                    voice_sessions = voice_sessions_result.scalars().all()
                    for vs in voice_sessions:
                        recordings_result = await db.execute(
                            select(VoiceRecording).where(VoiceRecording.voice_session_id == vs.id)
                        )
                        recordings = recordings_result.scalars().all()
                        deleted_voice_recordings += len(recordings)
                
                # RoundChoice 삭제 (room_id로 직접 삭제하는 대신 participant를 통해)
                participants_result = await db.execute(
                    select(RoomParticipant).where(RoomParticipant.room_id == room_id)
                )
                participants = participants_result.scalars().all()
                for p in participants:
                    await db.execute(
                        delete(RoundChoice).where(RoundChoice.participant_id == p.id)
                    )
                
                # ConsensusChoice 삭제
                await db.execute(delete(ConsensusChoice).where(ConsensusChoice.room_id == room_id))
                
                # VoiceRecording 삭제
                voice_sessions_result = await db.execute(
                    select(VoiceSession).where(VoiceSession.room_id == room_id)
                )
                voice_sessions = voice_sessions_result.scalars().all()
                for vs in voice_sessions:
                    await db.execute(delete(VoiceRecording).where(VoiceRecording.voice_session_id == vs.id))
                    await db.execute(delete(VoiceParticipant).where(VoiceParticipant.voice_session_id == vs.id))
                
                # VoiceSession 삭제
                await db.execute(delete(VoiceSession).where(VoiceSession.room_id == room_id))
                
                # RoomParticipant 삭제
                await db.execute(delete(RoomParticipant).where(RoomParticipant.room_id == room_id))
                
                # Room 삭제
                await db.delete(room)
                deleted_rooms += 1
    
    # User 삭제 (선택사항 - 조심해서 사용)
    if request.user_ids:
        for user_id in request.user_ids:
            # 먼저 해당 유저의 모든 참여 기록 삭제
            participations_result = await db.execute(
                select(RoomParticipant).where(RoomParticipant.user_id == user_id)
            )
            participations = participations_result.scalars().all()
            
            for participation in participations:
                await db.execute(
                    delete(RoundChoice).where(RoundChoice.participant_id == participation.id)
                )
            
            await db.execute(delete(RoomParticipant).where(RoomParticipant.user_id == user_id))
            
            # 음성 관련 삭제
            await db.execute(delete(VoiceRecording).where(VoiceRecording.user_id == user_id))
            await db.execute(delete(VoiceParticipant).where(VoiceParticipant.user_id == user_id))
            
            # User 삭제
            user_result = await db.execute(select(User).where(User.id == user_id))
            user = user_result.scalar_one_or_none()
            if user:
                await db.delete(user)
                deleted_users += 1
    
    await db.commit()
    
    return {
        "deleted_rooms": deleted_rooms,
        "deleted_users": deleted_users,
        "deleted_voice_recordings": deleted_voice_recordings,
        "message": f"Successfully deleted {deleted_rooms} rooms and {deleted_users} users"
    }


@router.get("/experiments/choices/analysis")
async def analyze_choices(
    topic: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    선택 데이터 분석
    - 라운드별 선택 분포
    - 역할별 선택 경향
    - 확신도 분석
    """
    # 라운드별 선택 분석 쿼리
    round_query = select(
        RoundChoice.round_number,
        RoundChoice.choice,
        func.count(RoundChoice.id).label('count'),
        func.avg(RoundChoice.confidence).label('avg_confidence')
    ).select_from(RoundChoice)
    
    if topic:
        # topic으로 필터링 (RoomParticipant를 통해 Room과 조인)
        round_query = round_query.join(
            RoomParticipant, RoundChoice.participant_id == RoomParticipant.id
        ).join(
            Room, RoomParticipant.room_id == Room.id
        ).where(Room.topic == topic)
    
    round_query = round_query.group_by(
        RoundChoice.round_number,
        RoundChoice.choice
    ).order_by(
        RoundChoice.round_number,
        RoundChoice.choice
    )
    
    round_result = await db.execute(round_query)
    round_analysis = round_result.all()
    
    # 역할별 선택 분석
    role_query = select(
        RoomParticipant.role_id,
        RoundChoice.choice,
        func.count(RoundChoice.id).label('count')
    ).select_from(RoundChoice).join(
        RoomParticipant,
        RoundChoice.participant_id == RoomParticipant.id
    )
    
    if topic:
        role_query = role_query.join(Room, RoomParticipant.room_id == Room.id).where(Room.topic == topic)
    
    role_query = role_query.group_by(
        RoomParticipant.role_id,
        RoundChoice.choice
    )
    
    role_result = await db.execute(role_query)
    role_analysis = role_result.all()
    
    return {
        "round_choices": [
            {
                "round_number": r.round_number,
                "choice": r.choice,
                "count": r.count,
                "avg_confidence": float(r.avg_confidence) if r.avg_confidence else None
            }
            for r in round_analysis
        ],
        "role_choices": [
            {
                "role_id": r.role_id,
                "choice": r.choice,
                "count": r.count
            }
            for r in role_analysis
        ]
    }


@router.get("/experiments/export/excel")
async def export_experiment_data_to_excel(
    started_only: bool = Query(False, description="시작된 게임만 포함"),
    with_consent_only: bool = Query(False, description="동의한 사용자만 포함"),
    topic: Optional[str] = Query(None, description="특정 주제 필터링"),
    db: AsyncSession = Depends(get_db)
):
    """
    실험 데이터를 엑셀 파일로 export
    
    형식:
    - episode_code: 게임 방 코드
    - participant_id: 참가자 ID
    - signup_date: 가입 날짜
    - username: 사용자 이름
    - email: 이메일
    - date_of_birth: 생년월일
    - gender: 성별
    - education_level: 교육 수준
    - R1~R5 각각: role, individual_choice, individual_confidence, group_choice, group_confidence
    """
    # 방 쿼리 (AsyncSession에 맞게 수정)
    stmt = select(Room)
    
    if started_only:
        stmt = stmt.where(Room.is_started == True)
    
    if topic:
        stmt = stmt.where(Room.topic == topic)
    
    result = await db.execute(stmt)
    rooms = result.scalars().all()
    
    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Experiment Data"
    
    # 헤더 스타일 설정
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 헤더 정의
    headers = [
        "episode_code", "participant_id", "signup_date", "username", "email",
        "date_of_birth", "gender", "education_level"
    ]
    
    # 라운드 1~5 헤더 추가
    for round_num in range(1, 6):
        headers.extend([
            f"R{round_num}_role",
            f"R{round_num}_individual_choice",
            f"R{round_num}_individual_confidence",
            f"R{round_num}_group_choice",
            f"R{round_num}_group_confidence"
        ])
    
    # 헤더 작성
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    current_row = 2
    
    for room in rooms:
        # 참가자 조회
        participant_stmt = select(RoomParticipant).where(RoomParticipant.room_id == room.id)
        participant_result = await db.execute(participant_stmt)
        participants = participant_result.scalars().all()
        
        # 라운드별 합의 선택 미리 조회
        consensus_stmt = select(ConsensusChoice).where(
            ConsensusChoice.room_id == room.id
        ).order_by(ConsensusChoice.round_number)
        
        consensus_result = await db.execute(consensus_stmt)
        consensus_choices = consensus_result.scalars().all()
        consensus_map = {cc.round_number: cc for cc in consensus_choices}
        
        for participant in participants:
            # 사용자 정보 조회
            user = None
            if participant.user_id:
                user_result = await db.get(User, participant.user_id)
                user = user_result
            
            # 동의 체크
            if with_consent_only and user:
                if not (user.data_consent and user.voice_consent):
                    continue
            
            # 라운드별 개인 선택 조회
            round_choice_stmt = select(RoundChoice).where(
                RoundChoice.participant_id == participant.id
            ).order_by(RoundChoice.round_number)
            
            round_choice_result = await db.execute(round_choice_stmt)
            round_choices = round_choice_result.scalars().all()
            round_choice_map = {rc.round_number: rc for rc in round_choices}
            
            # 기본 정보
            row_data = [
                room.room_code,  # episode_code
                participant.id,  # participant_id
                user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user else "",  # signup_date
                user.username if user else participant.nickname,  # username
                user.email if user else "",  # email
                user.birthdate if user else "",  # date_of_birth
                user.gender if user else "",  # gender
                user.education_level if user else ""  # education_level
            ]
            
            # 라운드 1~5 데이터 추가
            for round_num in range(1, 6):
                # 역할 정보 (모든 라운드에 동일하게 표시)
                role_str = _get_role_string(participant.role_id)
                
                # 개인 선택 정보
                individual_choice = ""
                individual_confidence = ""
                if round_num in round_choice_map:
                    rc = round_choice_map[round_num]
                    individual_choice = rc.choice if rc.choice is not None else ""
                    individual_confidence = rc.confidence if rc.confidence is not None else ""
                
                # 그룹 합의 정보
                group_choice = ""
                group_confidence = ""
                if round_num in consensus_map:
                    cc = consensus_map[round_num]
                    group_choice = cc.choice if cc.choice is not None else ""
                    group_confidence = cc.confidence if cc.confidence is not None else ""
                
                row_data.extend([
                    role_str,
                    individual_choice,
                    individual_confidence,
                    group_choice,
                    group_confidence
                ])
            
            # 행 작성
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
    
    # 컬럼 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 메모리에 엑셀 파일 저장
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 파일명 생성
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"experiment_data_{timestamp}.xlsx"
    
    # StreamingResponse로 반환
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _get_role_string(role_id: Optional[int]) -> str:
    """역할 ID를 문자열로 변환"""
    role_map = {
        1: "CG",   # Caregiver (요양보호사)
        2: "FAM",  # Family (가족)
        3: "DEV"   # Developer (AI 개발자)
    }
    return role_map.get(role_id, "") if role_id else ""


@router.get("/experiments/debug/counts")
async def get_data_counts(db: AsyncSession = Depends(get_db)):
    """
    디버깅용: 데이터베이스에 실제로 데이터가 몇 개나 있는지 확인
    """
    # Room 개수
    room_stmt = select(func.count()).select_from(Room)
    room_result = await db.execute(room_stmt)
    total_rooms = room_result.scalar()
    
    # RoomParticipant 개수
    participant_stmt = select(func.count()).select_from(RoomParticipant)
    participant_result = await db.execute(participant_stmt)
    total_participants = participant_result.scalar()
    
    # User 개수
    user_stmt = select(func.count()).select_from(User)
    user_result = await db.execute(user_stmt)
    total_users = user_result.scalar()
    
    # RoundChoice 개수
    round_choice_stmt = select(func.count()).select_from(RoundChoice)
    round_choice_result = await db.execute(round_choice_stmt)
    total_round_choices = round_choice_result.scalar()
    
    # ConsensusChoice 개수
    consensus_stmt = select(func.count()).select_from(ConsensusChoice)
    consensus_result = await db.execute(consensus_stmt)
    total_consensus = consensus_result.scalar()
    
    # 샘플 Room 데이터 (첫 5개)
    sample_rooms_stmt = select(Room).limit(5)
    sample_rooms_result = await db.execute(sample_rooms_stmt)
    sample_rooms = sample_rooms_result.scalars().all()
    
    return {
        "total_rooms": total_rooms,
        "total_participants": total_participants,
        "total_users": total_users,
        "total_round_choices": total_round_choices,
        "total_consensus_choices": total_consensus,
        "sample_rooms": [
            {
                "id": r.id,
                "room_code": r.room_code,
                "topic": r.topic,
                "is_started": r.is_started
            }
            for r in sample_rooms
        ]
    }


@router.get("/voice-recordings", response_model=VoiceRecordingsResponse)
async def get_voice_recordings(
    room_code: Optional[str] = Query(None, description="방 코드로 필터링 (6자리)"),
    room_id: Optional[int] = Query(None, description="방 ID로 필터링"),
    creator_id: Optional[int] = Query(None, description="방장(생성자) ID로 필터링"),
    user_id: Optional[int] = Query(None, description="특정 사용자 ID로 필터링"),
    start_date: Optional[datetime] = Query(None, description="시작 날짜 (YYYY-MM-DD)"),
    end_date: Optional[datetime] = Query(None, description="종료 날짜 (YYYY-MM-DD)"),
    skip: int = Query(0, ge=0, description="페이지네이션 오프셋"),
    limit: int = Query(100, ge=1, le=500, description="가져올 개수"),
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    음성 녹음 파일 목록 조회
    
    **필터 옵션:**
    - room_code: 방 입장 코드 (6자리)
    - room_id: 방 ID
    - creator_id: 방장(생성자) ID
    - user_id: 특정 사용자 ID
    - start_date/end_date: 날짜 범위
    
    **응답:**
    - 음성 녹음 파일의 S3 URL 목록
    - 방 정보, 사용자 정보 포함
    """
    # 기본 쿼리: VoiceRecording + VoiceSession + Room + User 조인
    query = select(VoiceRecording).join(
        VoiceSession, VoiceRecording.voice_session_id == VoiceSession.id
    ).join(
        Room, VoiceSession.room_id == Room.id
    ).outerjoin(
        User, VoiceRecording.user_id == User.id
    )
    
    # 필터 적용
    filters = []
    filters_applied = {}
    
    if room_code:
        filters.append(Room.room_code == room_code)
        filters_applied["room_code"] = room_code
    
    if room_id:
        filters.append(Room.id == room_id)
        filters_applied["room_id"] = room_id
    
    if creator_id:
        filters.append(Room.created_by == creator_id)
        filters_applied["creator_id"] = creator_id
    
    if user_id:
        filters.append(VoiceRecording.user_id == user_id)
        filters_applied["user_id"] = user_id
    
    if start_date:
        filters.append(VoiceRecording.created_at >= start_date)
        filters_applied["start_date"] = start_date.isoformat()
    
    if end_date:
        filters.append(VoiceRecording.created_at <= end_date)
        filters_applied["end_date"] = end_date.isoformat()
    
    if filters:
        query = query.where(and_(*filters))
    
    # 정렬: 최신순
    query = query.order_by(VoiceRecording.created_at.desc())
    
    # 총 개수 조회
    count_query = select(func.count()).select_from(VoiceRecording).join(
        VoiceSession, VoiceRecording.voice_session_id == VoiceSession.id
    ).join(
        Room, VoiceSession.room_id == Room.id
    )
    if filters:
        count_query = count_query.where(and_(*filters))
    
    total_count = (await db.execute(count_query)).scalar()
    
    # 페이지네이션 적용
    query = query.offset(skip).limit(limit)
    
    # 데이터 가져오기
    result = await db.execute(query.options(
        joinedload(VoiceRecording.voice_session).joinedload(VoiceSession.room),
        joinedload(VoiceRecording.user)
    ))
    recordings = result.scalars().unique().all()
    
    # 응답 데이터 구성
    recording_items = []
    for recording in recordings:
        room = recording.voice_session.room
        user = recording.user
        
        recording_items.append(VoiceRecordingItem(
            recording_id=recording.id,
            room_code=room.room_code,
            room_topic=room.topic,
            room_id=room.id,
            session_id=recording.voice_session.session_id,
            user_id=recording.user_id,
            guest_id=recording.guest_id,
            username=user.username if user else None,
            file_path=recording.file_path,  # S3 URL
            file_size=recording.file_size,
            duration=recording.duration,
            created_at=recording.created_at,
            is_processed=recording.is_processed
        ))
    
    return {
        "recordings": recording_items,
        "total_count": total_count,
        "page": skip // limit + 1,
        "page_size": limit,
        "filters_applied": filters_applied
    }
